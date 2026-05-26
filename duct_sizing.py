"""
Duct Sizing sheet writer for the HVAC loads scraper.

Builds a "Duct Sizing" sheet that mirrors the layout from
Load_Calculation_Document1.xlsm:

  Col A: row counter (hidden in the original, used by INDIRECT lookups
         in the original; here we just write integers since we have
         no Raw_Data sheet to look up against).
  Col B: Location (Zone name or "   Room <name>")
  Col C: Required CFM   (from DM HTML — "Required Supply CFM")
  Col D: Current  CFM   (the engineer edits this to match drawings)
  Col E: Defficiency formula (CHECK / numeric difference)
  Col F: Room-type tag formula (bath / rr or corridor / WIC / Corridor / "")
  F2:    Deficiency threshold (default 13)

The CHECK formula matches the workbook:
  - on Zone rows: CHECK if Current != Required
  - on Room rows: numeric deficiency if Current < Required by more than F2,
    but bath/rr/corridor/etc. rows are exempt (Col F not blank)

Usage:
    from openpyxl import Workbook
    from duct_sizing import write_duct_sizing

    wb = Workbook()
    write_duct_sizing(wb, report.supply_air)
    wb.save("output.xlsx")
"""

from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Threshold (in CFM) above which an under-supply on a room row counts
# as a deficiency. Bath/RR/corridor/WIC rows are exempt regardless of size.
DEFFICIENCY_THRESHOLD = 13

# Formulas — exactly mirror the workbook so engineers can edit
# Col D values and see Col E recompute the same way.
_E_FORMULA = (
    '=IF(ISNUMBER(SEARCH("zone",B{r})),'
        'IF(D{r}<>C{r},"CHECK",""),'
        'IF(F{r}="",'
            'IF(C{r}="","",IF(C{r}-D{r}>0,ABS(C{r}-D{r}),"")),'
            'IF(C{r}-D{r}>$F$2,ABS(C{r}-D{r}),"")'
        ')'
    ')'
)

_F_FORMULA = (
    '=IF(ISNUMBER(SEARCH("bath",B{r})),"bath",'
     'IF(ISNUMBER(SEARCH("rr",B{r})),"rr or corridor",'
     'IF(ISNUMBER(SEARCH("restroom",B{r})),"restroom",'
     'IF(ISNUMBER(SEARCH("toilet",B{r})),"toilet",'
     'IF(ISNUMBER(SEARCH("WIC",B{r})),"WIC",'
     'IF(ISNUMBER(SEARCH("corridor",B{r})),"Corridor",""))))))'
)


def _is_zone(location: str) -> bool:
    """SupplyAirRow locations are either 'Zone ...' or 'Room ...' after parsing."""
    return location.strip().lower().startswith("zone")


def write_duct_sizing(wb: Workbook, supply_air, sheet_name: str = "Duct Sizing"):
    """
    Add a Duct Sizing sheet to `wb` from a list[SupplyAirRow].

    Each Zone row gets a SUM() formula across its child Room rows so the
    Required and Current columns auto-roll-up. Room rows get the
    deficiency CHECK and room-type formulas in cols E and F.
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # --- Header rows (match Load_Calculation_Document1.xlsm exactly) ---
    ws["A1"] = ""                       # row counter starts at row 4 below
    ws["B1"] = "Supply Air Requirements"
    ws["B1"].font = Font(bold=True, size=14)

    ws["B2"] = "Location"
    ws["C2"] = "Required"
    ws["D2"] = "Current"
    ws["E2"] = "Defficiency"
    ws["F2"] = DEFFICIENCY_THRESHOLD     # threshold value referenced by Col E
    for col in ("B", "C", "D", "E"):
        ws[f"{col}2"].font = Font(bold=True)
        ws[f"{col}2"].alignment = Alignment(horizontal="center")

    ws["D3"] = "Supply CFM"
    ws["D3"].font = Font(italic=True, size=9)
    ws["D3"].alignment = Alignment(horizontal="center")

    # --- Body: walk the supply_air list and emit zone/room rows ---
    # We need to track each zone's row index so we can write SUM formulas
    # over its room range once we hit the next zone (or end of list).
    DATA_START = 5
    row = DATA_START
    zone_row: int | None = None         # row of the current zone header
    zone_first_room: int | None = None  # first room row under that zone

    bold = Font(bold=True)
    indent = Alignment(indent=0)
    room_indent = Alignment(indent=1)

    def _close_zone(end_row: int):
        """Write SUM formulas on the current zone's C and D cells covering
        zone_first_room..end_row, then reset the zone trackers."""
        nonlocal zone_row, zone_first_room
        if zone_row is None:
            return
        if zone_first_room is None or end_row < zone_first_room:
            # Zone with no rooms — leave the parsed values in place.
            zone_row = None
            zone_first_room = None
            return
        ws[f"C{zone_row}"] = f"=SUM(C{zone_first_room}:C{end_row})"
        ws[f"D{zone_row}"] = f"=SUM(D{zone_first_room}:D{end_row})"
        zone_row = None
        zone_first_room = None

    for sa in supply_air:
        loc = sa.location or ""
        is_zone = _is_zone(loc)

        # When we hit a new zone, close out the previous zone with SUMs
        if is_zone:
            _close_zone(row - 1)

        # Col A: simple row counter so engineers can see absolute row #
        ws[f"A{row}"] = row

        if is_zone:
            ws[f"B{row}"] = loc
            ws[f"B{row}"].font = bold
            # Required/Current placeholders — overwritten with SUM in _close_zone
            ws[f"C{row}"] = sa.required_supply_cfm or 0
            ws[f"D{row}"] = sa.current_supply_cfm or 0
            ws[f"C{row}"].font = bold
            ws[f"D{row}"].font = bold
            zone_row = row
            zone_first_room = None
        else:
            # Room row — indent the label visually like the original
            ws[f"B{row}"] = f"   Room {loc.replace('Room ', '', 1).strip()}"
            ws[f"C{row}"] = sa.required_supply_cfm or 0
            ws[f"D{row}"] = sa.current_supply_cfm or 0
            if zone_first_room is None and zone_row is not None:
                zone_first_room = row

        # E and F formulas go on every body row, zone or room
        ws[f"E{row}"] = _E_FORMULA.format(r=row)
        ws[f"F{row}"] = _F_FORMULA.format(r=row)

        row += 1

    # Close the final zone
    _close_zone(row - 1)

    # --- Column widths to match the original workbook ---
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].width = 37
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18

    # Freeze the header so the engineer can scroll the room list freely
    ws.freeze_panes = "A5"

    return ws


if __name__ == "__main__":
    # Smoke test: run Phase 1 then build a standalone Duct Sizing workbook.
    import runpy
    ns = runpy.run_path("phase1.py")
    wb = Workbook()
    # remove the default sheet
    del wb[wb.sheetnames[0]]
    write_duct_sizing(wb, ns["report"].supply_air)
    out = "duct_sizing_test.xlsx"
    wb.save(out)
    print(f"Wrote {out}")
