"""
HVAC Equipment Selector — Adicot Engineering
Uses equipment_db.xlsx to select Carrier split-system A/C or heat pump equipment.

Usage:
    python hvac_selector.py [--config config.json]
    python hvac_selector.py  (interactive mode)

Or import and call select_equipment() programmatically.
"""

import json
import math
import sys
import os
import argparse
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Optional DXF output ───────────────────────────────────────────────────────
try:
    import ezdxf
    HAS_EZDXF = True
except ImportError:
    HAS_EZDXF = False

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
DB_PATH = Path(__file__).parent / "equipment_db.xlsx"

# Equipment types
AC_SINGLE   = "GA5SAN5"
AC_TWO      = "GA8TAN5"
HP_SINGLE   = "GH5SAN5"
HP_TWO      = "GH8TAN5"

# ─────────────────────────────────────────────────────────────────────────────
# DATABASE LOADER
# ─────────────────────────────────────────────────────────────────────────────
_DB_CACHE = {}

def load_db(sheet: str) -> pd.DataFrame:
    if sheet not in _DB_CACHE:
        _DB_CACHE[sheet] = pd.read_excel(DB_PATH, sheet_name=sheet)
    return _DB_CACHE[sheet]

# ─────────────────────────────────────────────────────────────────────────────
# PSYCHROMETRIC HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def ewb_from_db_rh(db_f: float, rh_pct: float) -> float:
    """Approximate entering wet bulb from dry bulb (°F) and RH (%).
    Uses Magnus formula for dew point then iterates wet bulb.
    Accurate to ~0.3°F for typical HVAC conditions."""
    db_c = (db_f - 32) / 1.8
    rh = rh_pct / 100.0
    # Magnus approximation for dew point
    a, b = 17.625, 243.04
    alpha = math.log(rh) + a * db_c / (b + db_c)
    dp_c = b * alpha / (a - alpha)
    # Stull wet bulb approximation
    wb_c = (db_c * math.atan(0.151977 * (rh_pct + 8.313659) ** 0.5)
            + math.atan(db_c + rh_pct)
            - math.atan(rh_pct - 1.676331)
            + 0.00391838 * rh_pct ** 1.5 * math.atan(0.023101 * rh_pct)
            - 4.686035)
    return wb_c * 1.8 + 32

# ─────────────────────────────────────────────────────────────────────────────
# INTERPOLATION HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def interp2d(df: pd.DataFrame, col_x: str, x_val: float,
             col_y: str, y_val: float, col_out: str) -> float:
    """Bilinear interpolation from a table with two index columns."""
    xs = sorted(df[col_x].unique())
    ys = sorted(df[col_y].unique())
    x_lo = max([v for v in xs if v <= x_val], default=xs[0])
    x_hi = min([v for v in xs if v >= x_val], default=xs[-1])
    y_lo = max([v for v in ys if v <= y_val], default=ys[0])
    y_hi = min([v for v in ys if v >= y_val], default=ys[-1])

    def get(x, y):
        rows = df[(df[col_x] == x) & (df[col_y] == y)]
        if rows.empty:
            return None
        return float(rows[col_out].iloc[0])

    q11 = get(x_lo, y_lo); q12 = get(x_lo, y_hi)
    q21 = get(x_hi, y_lo); q22 = get(x_hi, y_hi)
    if None in (q11, q12, q21, q22):
        # Fall back to nearest
        val = get(x_lo, y_lo)
        return val if val is not None else float('nan')

    dx = (x_val - x_lo) / (x_hi - x_lo) if x_hi != x_lo else 0
    dy = (y_val - y_lo) / (y_hi - y_lo) if y_hi != y_lo else 0
    return (q11 * (1-dx) * (1-dy) + q21 * dx * (1-dy) +
            q12 * (1-dx) * dy     + q22 * dx * dy)


def interp1d_table(df: pd.DataFrame, col_x: str, x_val: float, col_out: str) -> float:
    xs = sorted(df[col_x].unique())
    x_lo = max([v for v in xs if v <= x_val], default=xs[0])
    x_hi = min([v for v in xs if v >= x_val], default=xs[-1])
    if x_lo == x_hi:
        row = df[df[col_x] == x_lo]
        return float(row[col_out].iloc[0])
    v_lo = float(df[df[col_x] == x_lo][col_out].iloc[0])
    v_hi = float(df[df[col_x] == x_hi][col_out].iloc[0])
    t = (x_val - x_lo) / (x_hi - x_lo)
    return v_lo + t * (v_hi - v_lo)

# ─────────────────────────────────────────────────────────────────────────────
# INDOOR COIL CAPACITY LOOKUP
# ─────────────────────────────────────────────────────────────────────────────
def get_indoor_coil_cap(series: str, fc_size: int, cfm: float,
                         ewb: float, edb: float, sst: float) -> tuple:
    """
    Returns (TC_corrected kBtu, SHC_corrected kBtu) for a given SST.
    SHC is corrected for EDB != 80°F.
    """
    sheet = "FG5_Cooling_Cap" if series == "FG5" else "FJ5_Cooling_Cap"
    df = load_db(sheet)
    sub = df[df["FC_Size"] == fc_size]

    # Snap CFM to nearest available
    cfms = sorted(sub["CFM"].unique())
    cfm_lo = max([c for c in cfms if c <= cfm], default=cfms[0])
    cfm_hi = min([c for c in cfms if c >= cfm], default=cfms[-1])

    # EWB snap
    ewbs = sorted(sub["EWB_F"].unique())
    ewb_lo = max([e for e in ewbs if e <= ewb], default=ewbs[0])
    ewb_hi = min([e for e in ewbs if e >= ewb], default=ewbs[-1])

    # SST snap
    ssts = sorted(sub["SST_F"].unique())
    sst_lo = max([s for s in ssts if s <= sst], default=ssts[0])
    sst_hi = min([s for s in ssts if s >= sst], default=ssts[-1])

    def get_vals(c, e, s):
        rows = sub[(sub["CFM"]==c) & (sub["EWB_F"]==e) & (sub["SST_F"]==s)]
        if rows.empty:
            return None, None, None
        r = rows.iloc[0]
        return float(r["TC_kBtu"]), float(r["SHC_kBtu"]), float(r["BF"])

    # Trilinear interpolation
    points = {}
    for c in [cfm_lo, cfm_hi]:
        for e in [ewb_lo, ewb_hi]:
            for s in [sst_lo, sst_hi]:
                tc, shc, bf = get_vals(c, e, s)
                if tc is not None:
                    points[(c,e,s)] = (tc, shc, bf)

    def tri_interp(key):
        tc11 = points.get((cfm_lo,ewb_lo,sst_lo),(None,None,None))[key]
        tc12 = points.get((cfm_lo,ewb_lo,sst_hi),(None,None,None))[key]
        tc21 = points.get((cfm_lo,ewb_hi,sst_lo),(None,None,None))[key]
        tc22 = points.get((cfm_lo,ewb_hi,sst_hi),(None,None,None))[key]
        tc31 = points.get((cfm_hi,ewb_lo,sst_lo),(None,None,None))[key]
        tc32 = points.get((cfm_hi,ewb_lo,sst_hi),(None,None,None))[key]
        tc41 = points.get((cfm_hi,ewb_hi,sst_lo),(None,None,None))[key]
        tc42 = points.get((cfm_hi,ewb_hi,sst_hi),(None,None,None))[key]
        vals = [v for v in [tc11,tc12,tc21,tc22,tc31,tc32,tc41,tc42] if v is not None]
        if not vals:
            return float('nan')
        dc = (cfm-cfm_lo)/(cfm_hi-cfm_lo) if cfm_hi!=cfm_lo else 0
        de = (ewb-ewb_lo)/(ewb_hi-ewb_lo) if ewb_hi!=ewb_lo else 0
        ds = (sst-sst_lo)/(sst_hi-sst_lo) if sst_hi!=sst_lo else 0
        # Trilinear
        c000 = tc11; c001 = tc12; c010 = tc21; c011 = tc22
        c100 = tc31; c101 = tc32; c110 = tc41; c111 = tc42
        for v in [c000,c001,c010,c011,c100,c101,c110,c111]:
            if v is None: return sum(vals)/len(vals)  # fallback
        return (c000*(1-dc)*(1-de)*(1-ds) + c001*(1-dc)*(1-de)*ds +
                c010*(1-dc)*de*(1-ds)     + c011*(1-dc)*de*ds +
                c100*dc*(1-de)*(1-ds)     + c101*dc*(1-de)*ds +
                c110*dc*de*(1-ds)         + c111*dc*de*ds)

    tc = tri_interp(0)
    shc = tri_interp(1)
    bf = tri_interp(2)

    # Correct SHC for EDB != 80°F
    if not math.isnan(shc) and not math.isnan(bf):
        cf = 1.09 * (1 - bf) * (edb - 80) * cfm / 1000.0
        print(cf)
        shc_corr = shc + cf  # positive when EDB>80, negative when EDB<80
    else:
        shc_corr = shc

    return tc, shc_corr


# ─────────────────────────────────────────────────────────────────────────────
# CONDENSING UNIT CAPACITY LOOKUP (A/C)
# ─────────────────────────────────────────────────────────────────────────────
def get_condenser_cap(odu_model: str, series: str, stage: str,
                       sst: float, odb: float) -> tuple:
    """Returns (TCG kBtu, SDT °F, kW_ODU) for given SST and ODB."""
    sheet = "GA5_Condenser_Only" if series == AC_SINGLE else "GA8_Condenser_Only"
    df = load_db(sheet)
    sub = df[df["Model"] == odu_model]
    if series == AC_TWO:
        sub = sub[sub["Stage"] == stage]

    tcg = interp2d(sub, "SST_F", sst, "ODB_F", odb, "TCG_kBtu")
    sdt = interp2d(sub, "SST_F", sst, "ODB_F", odb, "SDT_F")
    kw  = interp2d(sub, "SST_F", sst, "ODB_F", odb, "kW_ODU")
    return tcg, sdt, kw


# ─────────────────────────────────────────────────────────────────────────────
# SYSTEM OPERATING POINT (A/C) — iterative SST solver
# ─────────────────────────────────────────────────────────────────────────────
def find_ac_operating_point(odu_model: str, odu_series: str, fc_series: str,
                             fc_size: int, cfm: float, ewb: float, edb: float,
                             odb: float, stage: str = "High") -> dict:
    """
    Find SST where indoor coil TC intersects condenser TCG.
    Returns dict with tc, shc, sst, sdt, kw_odu.
    """
    best = None
    min_diff = float('inf')

    for sst in [30, 32, 34, 36, 38, 40, 42, 44, 46, 48, 50, 52, 54]:
        tc_indoor, shc = get_indoor_coil_cap(fc_series, fc_size, cfm, ewb, edb, sst)
        tcg_cond, sdt, kw = get_condenser_cap(odu_model, odu_series, stage, sst, odb)

        if math.isnan(tc_indoor) or math.isnan(tcg_cond):
            continue

        diff = abs(tc_indoor - tcg_cond)
        if diff < min_diff:
            min_diff = diff
            best = {
                "sst": sst, "tc": tc_indoor, "shc": shc,
                "tcg_cond": tcg_cond, "sdt": sdt, "kw_odu": kw,
                "balance_diff": diff
            }

    # Refine with finer step around best SST
    if best:
        sst_range = [best["sst"] + d * 0.5 for d in range(-4, 5)]
        for sst in sst_range:
            tc_indoor, shc = get_indoor_coil_cap(fc_series, fc_size, cfm, ewb, edb, sst)
            tcg_cond, sdt, kw = get_condenser_cap(odu_model, odu_series, stage, sst, odb)
            if math.isnan(tc_indoor) or math.isnan(tcg_cond):
                continue
            diff = abs(tc_indoor - tcg_cond)
            if diff < min_diff:
                min_diff = diff
                best = {"sst": sst, "tc": tc_indoor, "shc": shc,
                        "tcg_cond": tcg_cond, "sdt": sdt, "kw_odu": kw,
                        "balance_diff": diff}
    return best or {}


# ─────────────────────────────────────────────────────────────────────────────
# HEAT PUMP COOLING CAPACITY
# ─────────────────────────────────────────────────────────────────────────────
def get_hp_cooling_cap(odu_model: str, series: str, fc_size: int,
                        cfm: float, ewb: float, edb: float,
                        odb: float, stage: str = "High") -> tuple:
    """Returns (TC_net kBtu, Sens_net kBtu, sys_kW) for heat pump cooling."""
    sheet = "GH5_Detail_Cool" if series == HP_SINGLE else "GH8_Detail_Cool"
    df = load_db(sheet)
    sub = df[df["ODU_Model"] == odu_model]
    if series == HP_TWO:
        sub = sub[sub["Stage"] == stage]

    if sub.empty:
        return float('nan'), float('nan'), float('nan')

    # Snap CFM
    cfms = sorted(sub["CFM"].unique())
    cfm_lo = max([c for c in cfms if c <= cfm], default=cfms[0])
    cfm_hi = min([c for c in cfms if c >= cfm], default=cfms[-1])

    # EWB snap
    ewbs = sorted(sub["EWB_F"].unique())
    ewb_lo = max([e for e in ewbs if e <= ewb], default=ewbs[0])
    ewb_hi = min([e for e in ewbs if e >= ewb], default=ewbs[-1])

    # ODB snap
    odbs = sorted(sub["ODB_F"].unique())
    odb_lo = max([o for o in odbs if o <= odb], default=odbs[0])
    odb_hi = min([o for o in odbs if o >= odb], default=odbs[-1])

    def get_row(c, e, o):
        rows = sub[(sub["CFM"]==c) & (sub["EWB_F"]==e) & (sub["ODB_F"]==o)]
        if rows.empty: return None, None, None
        r = rows.iloc[0]
        return float(r["TC_Net_kBtu"]), float(r["Sens_Net_kBtu"]), float(r["Sys_kW"])

    def lerp(a, b, t): return a + (b - a) * t if a is not None and b is not None else (a or b)

    dc = (cfm-cfm_lo)/(cfm_hi-cfm_lo) if cfm_hi!=cfm_lo else 0
    de = (ewb-ewb_lo)/(ewb_hi-ewb_lo) if ewb_hi!=ewb_lo else 0
    do = (odb-odb_lo)/(odb_hi-odb_lo) if odb_hi!=odb_lo else 0

    results = {}
    for idx, col in enumerate(["tc","sens","kw"]):
        vals = []
        for c in [cfm_lo, cfm_hi]:
            for e in [ewb_lo, ewb_hi]:
                for o in [odb_lo, odb_hi]:
                    v = get_row(c,e,o)[idx]
                    vals.append(v)
        # Trilinear
        c000,c001,c010,c011,c100,c101,c110,c111 = vals
        good = all(v is not None for v in vals)
        if good:
            results[col] = (c000*(1-dc)*(1-de)*(1-do) + c001*(1-dc)*(1-de)*do +
                            c010*(1-dc)*de*(1-do) + c011*(1-dc)*de*do +
                            c100*dc*(1-de)*(1-do) + c101*dc*(1-de)*do +
                            c110*dc*de*(1-do)     + c111*dc*de*do)
        else:
            good_vals = [v for v in vals if v is not None]
            results[col] = sum(good_vals)/len(good_vals) if good_vals else float('nan')

    # Correct sensible for EDB != 80°F (835 Btuh per 1000 CFM per °F)
    sens_corr = results["sens"] + 0.835 * cfm/1000 * (edb - 80)

    return results["tc"], sens_corr, results["kw"]


# ─────────────────────────────────────────────────────────────────────────────
# HEAT PUMP HEATING CAPACITY
# ─────────────────────────────────────────────────────────────────────────────
def get_hp_heating_cap(odu_model: str, series: str,
                        cfm: float, edb: float, odb: float,
                        stage: str = "High") -> tuple:
    """Returns (Cap_Total kBtu, Cap_Integ kBtu, Sys_kW) for heat pump heating."""
    sheet = "GH5_Heating" if series == HP_SINGLE else "GH8_Heating"
    df = load_db(sheet)
    sub = df[df["ODU_Model"] == odu_model]
    if series == HP_TWO:
        sub = sub[sub["Stage"] == stage]

    if sub.empty:
        return float('nan'), float('nan'), float('nan')

    cfms = sorted(sub["CFM"].unique())
    cfm_use = min(cfms, key=lambda c: abs(c - cfm))

    edbs = sorted(sub["EDB_F"].unique())
    edb_lo = max([e for e in edbs if e <= edb], default=edbs[0])
    edb_hi = min([e for e in edbs if e >= edb], default=edbs[-1])

    odbs = sorted(sub["ODB_F"].unique())
    odb_lo = max([o for o in odbs if o <= odb], default=odbs[0])
    odb_hi = min([o for o in odbs if o >= odb], default=odbs[-1])

    def get_row(e, o):
        rows = sub[(sub["EDB_F"]==e) & (sub["CFM"]==cfm_use) & (sub["ODB_F"]==o)]
        if rows.empty: return None, None, None
        r = rows.iloc[0]
        return float(r["Cap_Total_kBtu"]), float(r["Cap_Integ_kBtu"]), float(r["Sys_kW"])

    de = (edb-edb_lo)/(edb_hi-edb_lo) if edb_hi!=edb_lo else 0
    do = (odb-odb_lo)/(odb_hi-odb_lo) if odb_hi!=odb_lo else 0

    t00 = get_row(edb_lo, odb_lo); t01 = get_row(edb_lo, odb_hi)
    t10 = get_row(edb_hi, odb_lo); t11 = get_row(edb_hi, odb_hi)

    results = []
    for i in range(3):
        v00, v01, v10, v11 = t00[i], t01[i], t10[i], t11[i]
        if all(v is not None for v in [v00,v01,v10,v11]):
            val = (v00*(1-de)*(1-do) + v01*(1-de)*do +
                   v10*de*(1-do)     + v11*de*do)
        else:
            vals = [v for v in [v00,v01,v10,v11] if v is not None]
            val = sum(vals)/len(vals) if vals else float('nan')
        results.append(val)

    return tuple(results)


# ─────────────────────────────────────────────────────────────────────────────
# CFM SELECTION
# ─────────────────────────────────────────────────────────────────────────────
def get_nominal_cfm(fc_series: str, fc_size: int, odu_kbtu: int = None) -> float:
    """Get nominal cooling CFM for a fan coil."""
    if fc_series == "FG5":
        df = load_db("FG5_Airflow")
        sub = df[df["FC_Size"] == fc_size]
        if odu_kbtu is not None:
            row = sub[sub["ODU_Size_kBtu"] == odu_kbtu]
            if not row.empty:
                return float(row["A/C_Cool_Nom_CFM"].iloc[0])  # always max (nominal) cooling CFM for FG5
        if not sub.empty:
            return float(sub["A/C_Cool_Nom_CFM"].iloc[-1])  # largest match
    else:  # FJ5
        df = load_db("FJ5_Airflow")
        row = df[df["FC_Size"] == fc_size]
        if not row.empty:
            return float(row["Nom_CFM"].iloc[0])
    return fc_size * 400 / 12  # rough fallback: 400 CFM/ton


# ─────────────────────────────────────────────────────────────────────────────
# MAIN SELECTION FUNCTION
# ─────────────────────────────────────────────────────────────────────────────
def select_equipment(zones: list, outdoor_db: float, outdoor_wb: float,
                     entering_db: float = 80.0, entering_ewb: float = 67.0,
                     cap_min_pct: float = 100.0, cap_max_pct: float = 115.0,
                     mode: str = "cooling",
                     equipment_types: list = None) -> list:
    """
    Select equipment for each zone.

    Args:
        zones: list of dicts with keys:
            - name (str): zone identifier
            - total_cooling_kbtu (float): design cooling load
            - sensible_cooling_kbtu (float): design sensible load
            - total_heating_kbtu (float, optional): design heating load
        outdoor_db: outdoor design dry bulb °F
        outdoor_wb: outdoor design wet bulb °F
        entering_db: entering air dry bulb °F (default 80)
        entering_ewb: entering air wet bulb °F (default 67)
        cap_min_pct: minimum capacity as % of load (default 100)
        cap_max_pct: maximum capacity as % of load (default 115)
        mode: "cooling" or "heating" (for priority; always checks cooling)
        equipment_types: list of series to consider, e.g. ["GA5SAN5","GH5SAN5"]
                         None = all types

    Returns:
        list of selection result dicts, one per zone
    """
    if equipment_types is None:
        equipment_types = [AC_SINGLE, AC_TWO, HP_SINGLE, HP_TWO]

    edb = entering_db
    indoor_ewb = entering_ewb

    combos_df = load_db("Valid_Combinations")

    results = []

    for zone in zones:
        zone_name = zone.get("name", "Zone")
        tc_load = zone["total_cooling_kbtu"]
        shc_load = zone["sensible_cooling_kbtu"]
        htg_load = zone.get("total_heating_kbtu", 0)

        tc_min = tc_load * cap_min_pct / 100.0
        tc_max = tc_load * cap_max_pct / 100.0

        candidates = []
        _oob_candidates = []

        for _, combo in combos_df.iterrows():
            odu_series = combo["ODU_Series"]
            if odu_series not in equipment_types:
                continue

            nom_kbtu = int(combo["Nom_kBtu"])
            odu_model = combo["ODU_Model"]
            fc_series = combo["IDU_Series"]
            nom_tons = combo["Nom_Tons"]
            stage = "High"  # always use High for peak cooling selection

            # Get nominal CFM
            cfm = get_nominal_cfm(fc_series, nom_kbtu, nom_kbtu)

            # ── Calculate cooling capacity ────────────────────────────────
            if odu_series in [AC_SINGLE, AC_TWO]:
                # A/C: solve for SST balance point
                op = find_ac_operating_point(
                    odu_model, odu_series, fc_series,
                    nom_kbtu, cfm, indoor_ewb, edb, outdoor_db, stage
                )
                if not op:
                    continue
                tc_sys = op["tc"]
                shc_sys = op["shc"]
                kw_sys = op.get("kw_odu", None)
                sst = op["sst"]
                sdt = op.get("sdt", None)
            else:
                # HP: direct table lookup
                tc_sys, shc_sys, kw_sys = get_hp_cooling_cap(
                    odu_model, odu_series, nom_kbtu,
                    cfm, indoor_ewb, edb, outdoor_db, stage
                )
                sst = None; sdt = None

            if math.isnan(tc_sys) or math.isnan(shc_sys):
                continue

            # ── Capacity check ────────────────────────────────────────────
            out_of_bounds = tc_sys < tc_min or tc_sys > tc_max
            sensible_fail = shc_sys < shc_load
            if out_of_bounds or sensible_fail:
                if not sensible_fail:  # track for fallback if only TC is out of bounds
                    _oob_candidates.append({
                        'tc_sys': tc_sys, 'shc_sys': shc_sys,
                        'nom_kbtu': nom_kbtu, 'odu_model': odu_model,
                        'odu_series': odu_series, 'nom_tons': nom_tons,
                        'fc_series': fc_series, 'cfm': cfm,
                        'idu_pattern': combo['IDU_Model_Pattern'],
                        'sst': sst, 'sdt': sdt, 'kw_sys': kw_sys,
                        'undershoot': tc_sys < tc_min,
                    })
                continue

            # ── Heating capacity (heat pumps only) ────────────────────────
            htg_cap = None
            htg_kw = None
            heating_ok = True
            if odu_series in [HP_SINGLE, HP_TWO] and htg_load > 0:
                htg_total, htg_integ, htg_kw_sys = get_hp_heating_cap(
                    odu_model, odu_series, cfm,
                    entering_db, outdoor_db, stage
                )
                htg_cap = htg_integ  # use integrated (defrost-corrected)
                htg_kw = htg_kw_sys
                # Flag if heating capacity < load (supplemental heat needed)
                if not math.isnan(htg_cap) and htg_cap < htg_load:
                    heating_ok = False  # still include but flag

            # ── Efficiency (EER at these conditions) ─────────────────────
            if kw_sys and not math.isnan(kw_sys) and kw_sys > 0:
                eer = (tc_sys * 1000) / (kw_sys * 1000)  # Btuh / W
            else:
                eer = None

            candidates.append({
                "zone": zone_name,
                "odu_model": odu_model,
                "odu_series": odu_series,
                "nom_tons": nom_tons,
                "nom_kbtu": nom_kbtu,
                "fc_series": fc_series,
                "idu_pattern": combo["IDU_Model_Pattern"],
                "cfm": cfm,
                "tc_sys_kbtu": round(tc_sys, 2),
                "shc_sys_kbtu": round(shc_sys, 2),
                "tc_load_kbtu": tc_load,
                "shc_load_kbtu": shc_load,
                "tc_pct": round(tc_sys / tc_load * 100, 1),
                "sst_f": round(sst, 1) if sst else None,
                "sdt_f": round(sdt, 1) if sdt else None,
                "kw_odu": round(kw_sys, 2) if kw_sys and not math.isnan(kw_sys) else None,
                "eer_site": round(eer, 2) if eer else None,
                "htg_cap_kbtu": round(htg_cap, 2) if htg_cap and not math.isnan(htg_cap) else None,
                "htg_load_kbtu": htg_load,
                "htg_ok": heating_ok,
                "outdoor_db": outdoor_db,
                "outdoor_wb": outdoor_wb,
                "entering_ewb": round(indoor_ewb, 1),
            })

        # ── Sort: prefer smallest unit that meets load, then by EER ──────
        candidates.sort(key=lambda x: (x["nom_kbtu"], -(x["eer_site"] or 0)))

        # ── If no in-bounds candidates, find closest smaller and larger ──
        next_smaller = next_larger = None
        if not candidates and _oob_candidates:
            undershoots = [c for c in _oob_candidates if c['undershoot']]
            overshoots  = [c for c in _oob_candidates if not c['undershoot']]
            if undershoots:
                # Largest unit still below min bound
                next_smaller = max(undershoots, key=lambda c: c['tc_sys'])
            if overshoots:
                # Smallest unit above max bound
                next_larger = min(overshoots, key=lambda c: c['tc_sys'])

        results.append({
            "zone": zone_name,
            "tc_load": tc_load,
            "shc_load": shc_load,
            "selected": candidates[0] if candidates else None,
            "next_smaller": next_smaller,
            "next_larger": next_larger,
            "all_candidates": candidates,
        })

    return results


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL SCHEDULE OUTPUT
# ─────────────────────────────────────────────────────────────────────────────
def write_excel_schedule(results: list, filepath: str,
                          project_name: str = "PROJECT",
                          outdoor_db: float = None, outdoor_wb: float = None,
                          cap_min_pct: float = 100, cap_max_pct: float = 115):
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Schedule"

    # Styles
    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", size=10)
    sub_fill = PatternFill("solid", start_color="D6E4F0")
    alt_fill = PatternFill("solid", start_color="EBF3FB")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = h_font; c.fill = h_fill; c.alignment = center; c.border = bdr
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width

    def dcell(row, col, val, alt=False, bold=False, center_align=True):
        c = ws.cell(row=row, column=col, value=val)
        if alt: c.fill = alt_fill
        if bold: c.font = Font(bold=True)
        c.alignment = center if center_align else left
        c.border = bdr

    # Title
    ws.merge_cells("A1:R1")
    tc = ws["A1"]
    tc.value = f"MECHANICAL — AC/HP EQUIPMENT SCHEDULE — {project_name}"
    tc.font = Font(bold=True, size=13, color="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Design conditions row
    ws.merge_cells("A2:R2")
    cond_str = ""
    if outdoor_db: cond_str += f"  Outdoor DB: {outdoor_db}°F"
    if outdoor_wb: cond_str += f"  |  Outdoor WB: {outdoor_wb}°F"
    cond_str += "  |  Capacities at design conditions (corrected)"
    ws["A2"].value = cond_str
    ws["A2"].font = Font(italic=True, size=9, color="606060")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Column headers — row 3
    columns = [
        ("Tag / Zone", 14), ("Outdoor Unit Model", 22), ("Series", 10),
        ("Nom. Tons", 10), ("Indoor Unit", 22), ("Fan Coil\nSeries", 10),
        ("Nom. CFM", 10),
        ("Design TC\n(kBtu/h)", 12), ("Sys TC\n(kBtu/h)", 12), ("TC\n(%)", 8),
        ("Design SHC\n(kBtu/h)", 13), ("Sys SHC\n(kBtu/h)", 13),
        ("Op. SST\n(°F)", 10), ("Op. SDT\n(°F)", 10),
        ("ODU kW\n@ Design", 11), ("Site EER\n(Btu/Wh)", 10),
        ("HP Htg Cap\n(kBtu/h)", 13), ("Htg OK?", 8),
    ]
    ws.row_dimensions[3].height = 36
    for col_idx, (hdr, width) in enumerate(columns, 1):
        hcell(3, col_idx, hdr, width)

    row = 4
    for i, res in enumerate(results):
        sel = res["selected"]
        alt = (i % 2 == 0)
        if sel is None:
            dcell(row, 1, res["zone"], alt); dcell(row, 2, "NO MATCH FOUND", alt, bold=True)
            for c in range(3, len(columns)+1): dcell(row, c, "—", alt)
            row += 1
            continue

        vals = [
            sel["zone"],
            sel["odu_model"],
            sel["odu_series"],
            f"{sel['nom_tons']}T",
            sel["idu_pattern"],
            sel["fc_series"],
            int(sel["cfm"]),
            sel["tc_load_kbtu"],
            sel["tc_sys_kbtu"],
            f"{sel['tc_pct']}%",
            sel["shc_load_kbtu"],
            sel["shc_sys_kbtu"],
            sel["sst_f"] or "—",
            sel["sdt_f"] or "—",
            sel["kw_odu"] or "—",
            sel["eer_site"] or "—",
            sel["htg_cap_kbtu"] or "N/A",
            "YES" if sel["htg_ok"] else "SUPPL. HTG REQ.",
        ]
        for col_idx, val in enumerate(vals, 1):
            dcell(row, col_idx, val, alt,
                  bold=(col_idx == 2),
                  center_align=(col_idx != 5))
        row += 1

    # Notes section
    row += 1
    ws.cell(row=row, column=1, value="NOTES:").font = Font(bold=True)
    notes = [
        "1. All capacities shown are corrected to actual design outdoor and indoor conditions.",
        "2. Sensible Heat Capacity (SHC) corrected for entering dry-bulb ≠ 80°F per manufacturer correction formula.",
        "3. HP Heating Capacity shown as integrated value with defrost penalty subtracted per AHRI 210/240.",
        "4. 'SUPPL. HTG REQ.' indicates heat pump heating capacity < design heating load; supplemental electric heat required.",
        "5. Op. SST = Operating Saturated Suction Temperature (A/C systems only — iterative balance point).",
        "6. Verify AHRI certified combinations at ahridirectory.org before finalizing selections.",
        f"7. Selection range: {cap_min_pct:.0f}% – {cap_max_pct:.0f}% of design cooling load.",
        "8. Source data: Carrier FG5, FJ5, GA5SAN5, GA8TAN5, GH5SAN5, GH8TAN5 Product Data. All Puron Advance™ (R-454B).",
    ]
    for note in notes:
        row += 1
        ws.cell(row=row, column=1, value=note).font = Font(size=9)

    wb.save(filepath)
    print(f"Schedule saved: {filepath}")


# ─────────────────────────────────────────────────────────────────────────────
# DXF SCHEDULE OUTPUT
# ─────────────────────────────────────────────────────────────────────────────
def write_dxf_schedule(results: list, filepath: str,
                        project_name: str = "PROJECT",
                        outdoor_db: float = None, outdoor_wb: float = None):
    if not HAS_EZDXF:
        print("ezdxf not installed — skipping DXF output. Run: pip install ezdxf")
        return

    doc = ezdxf.new("R2010")
    doc.header["$INSUNITS"] = 1  # inches

    # Layers
    doc.layers.add("SCHEDULE_BORDER", color=7)
    doc.layers.add("SCHEDULE_HEADER", color=5)
    doc.layers.add("SCHEDULE_TEXT",   color=7)
    doc.layers.add("SCHEDULE_TITLE",  color=1)
    doc.layers.add("SCHEDULE_GRID",   color=8)

    msp = doc.modelspace()

    # ── Layout parameters ─────────────────────────────────────────────────
    # Column widths in inches
    col_widths = [2.0, 3.5, 1.6, 1.2, 3.5, 1.2, 1.2,
                  1.4, 1.4, 0.9, 1.5, 1.5,
                  1.2, 1.2, 1.3, 1.2, 1.5, 1.4]
    col_headers = [
        "TAG / ZONE", "OUTDOOR UNIT MODEL", "SERIES", "TONS", "INDOOR UNIT",
        "FC SER.", "CFM",
        "DESGN TC\n(kBtu/h)", "SYS TC\n(kBtu/h)", "TC %",
        "DESGN SHC\n(kBtu/h)", "SYS SHC\n(kBtu/h)",
        "SST °F", "SDT °F", "ODU kW", "EER", "HP HTG\n(kBtu/h)", "HTG OK",
    ]
    row_height = 0.40    # inches per data row
    hdr_height = 0.55    # header row height
    title_height = 0.50
    txt_size = 0.09
    hdr_txt_size = 0.08
    title_txt_size = 0.14

    table_w = sum(col_widths)
    n_rows = len(results) + 1  # header + data
    table_h = title_height + hdr_height + row_height * len(results) + 0.5  # notes gap

    ox = 0.0  # origin x
    oy = 0.0  # origin y (bottom)
    y_top = oy + table_h

    # ── Title bar ─────────────────────────────────────────────────────────
    title_y_bot = y_top - title_height
    msp.add_lwpolyline(
        [(ox, title_y_bot), (ox+table_w, title_y_bot),
         (ox+table_w, y_top), (ox, y_top), (ox, title_y_bot)],
        close=True, dxfattribs={"layer": "SCHEDULE_BORDER", "lineweight": 50}
    )
    cond_str = ""
    if outdoor_db: cond_str = f"  ODB: {outdoor_db}°F"
    if outdoor_wb: cond_str += f"  OWB: {outdoor_wb}°F"
    title_text = f"MECHANICAL — AC/HP EQUIPMENT SCHEDULE — {project_name}{cond_str}"
    msp.add_text(title_text, dxfattribs={
        "layer": "SCHEDULE_TITLE", "height": title_txt_size,
        "insert": (ox + 0.1, (y_top + title_y_bot) / 2 - title_txt_size/2)
    })

    # ── Header row ────────────────────────────────────────────────────────
    hdr_y_bot = title_y_bot - hdr_height
    msp.add_lwpolyline(
        [(ox, hdr_y_bot), (ox+table_w, hdr_y_bot),
         (ox+table_w, title_y_bot), (ox, title_y_bot), (ox, hdr_y_bot)],
        close=True, dxfattribs={"layer": "SCHEDULE_HEADER", "lineweight": 35}
    )
    # Hatch header
    hatch = msp.add_hatch(color=5, dxfattribs={"layer": "SCHEDULE_HEADER"})
    hatch.paths.add_polyline_path(
        [(ox, hdr_y_bot), (ox+table_w, hdr_y_bot),
         (ox+table_w, title_y_bot), (ox, title_y_bot)], is_closed=True
    )
    hatch.set_solid_fill()
    hatch.dxf.color = 9  # light grey

    x = ox
    for col_i, (cw, ch) in enumerate(zip(col_widths, col_headers)):
        # Vertical grid line
        msp.add_line((x, hdr_y_bot), (x, title_y_bot),
                     dxfattribs={"layer": "SCHEDULE_GRID"})
        # Header text (first line of multiline)
        first_line = ch.split("\n")[0]
        msp.add_text(first_line, dxfattribs={
            "layer": "SCHEDULE_HEADER",
            "height": hdr_txt_size, "color": 7,
            "insert": (x + 0.03, (title_y_bot + hdr_y_bot)/2 + hdr_txt_size*0.3)
        })
        if "\n" in ch:
            second_line = ch.split("\n")[1]
            msp.add_text(second_line, dxfattribs={
                "layer": "SCHEDULE_HEADER",
                "height": hdr_txt_size, "color": 7,
                "insert": (x + 0.03, (title_y_bot + hdr_y_bot)/2 - hdr_txt_size*1.1)
            })
        x += cw
    # Last vertical line
    msp.add_line((x, hdr_y_bot), (x, title_y_bot),
                 dxfattribs={"layer": "SCHEDULE_GRID"})

    # ── Data rows ─────────────────────────────────────────────────────────
    row_y_top = hdr_y_bot
    for i, res in enumerate(results):
        row_y_bot = row_y_top - row_height
        sel = res["selected"]

        # Row background (alternate)
        if i % 2 == 1:
            hatch_row = msp.add_hatch(dxfattribs={"layer": "SCHEDULE_GRID"})
            hatch_row.paths.add_polyline_path(
                [(ox, row_y_bot), (ox+table_w, row_y_bot),
                 (ox+table_w, row_y_top), (ox, row_y_top)], is_closed=True
            )
            hatch_row.set_solid_fill()
            hatch_row.dxf.color = 253  # very light grey

        # Row border
        msp.add_line((ox, row_y_bot), (ox+table_w, row_y_bot),
                     dxfattribs={"layer": "SCHEDULE_GRID"})

        if sel:
            row_vals = [
                sel["zone"],
                sel["odu_model"],
                sel["odu_series"],
                f"{sel['nom_tons']}T",
                sel["idu_pattern"].replace("*", ""),
                sel["fc_series"],
                str(int(sel["cfm"])),
                str(sel["tc_load_kbtu"]),
                str(sel["tc_sys_kbtu"]),
                f"{sel['tc_pct']}%",
                str(sel["shc_load_kbtu"]),
                str(sel["shc_sys_kbtu"]),
                str(sel["sst_f"]) if sel["sst_f"] else "—",
                str(sel["sdt_f"]) if sel["sdt_f"] else "—",
                str(sel["kw_odu"]) if sel["kw_odu"] else "—",
                str(sel["eer_site"]) if sel["eer_site"] else "—",
                str(sel["htg_cap_kbtu"]) if sel["htg_cap_kbtu"] else "N/A",
                "OK" if sel["htg_ok"] else "SUPPL",
            ]
        else:
            row_vals = [res["zone"], "NO MATCH"] + ["—"] * (len(col_widths)-2)

        x = ox
        for col_i, (cw, val) in enumerate(zip(col_widths, row_vals)):
            msp.add_line((x, row_y_bot), (x, row_y_top),
                         dxfattribs={"layer": "SCHEDULE_GRID"})
            # Clip text to column width
            max_chars = max(1, int(cw / txt_size * 1.4))
            display_val = str(val)[:max_chars]
            msp.add_text(display_val, dxfattribs={
                "layer": "SCHEDULE_TEXT",
                "height": txt_size,
                "insert": (x + 0.04, row_y_bot + row_height*0.28)
            })
            x += cw
        msp.add_line((x, row_y_bot), (x, row_y_top),
                     dxfattribs={"layer": "SCHEDULE_GRID"})

        row_y_top = row_y_bot

    # Outer border
    msp.add_lwpolyline(
        [(ox, row_y_top), (ox+table_w, row_y_top),
         (ox+table_w, y_top), (ox, y_top), (ox, row_y_top)],
        close=True, dxfattribs={"layer": "SCHEDULE_BORDER", "lineweight": 70}
    )

    # Notes
    note_y = row_y_top - 0.25
    notes_text = [
        "NOTES: 1. Capacities corrected to design conditions.  2. SHC corrected for EDB ≠ 80°F.  3. HP Htg = integrated defrost-corrected value.",
        "4. 'SUPPL' = supplemental heat required.  5. Verify AHRI certified combinations at ahridirectory.org.  6. All units: Puron Advance™ (R-454B)."
    ]
    for nt in notes_text:
        msp.add_text(nt, dxfattribs={"layer": "SCHEDULE_TEXT", "height": 0.07,
                                      "insert": (ox, note_y)})
        note_y -= 0.15

    doc.saveas(filepath)
    print(f"DXF schedule saved: {filepath}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI / INTERACTIVE
# ─────────────────────────────────────────────────────────────────────────────
def run_interactive():
    print("\n" + "="*60)
    print("  ADICOT ENGINEERING — HVAC EQUIPMENT SELECTOR")
    print("  Carrier Puron Advance™ (R-454B) Split Systems")
    print("="*60 + "\n")

    project = input("Project name: ").strip() or "PROJECT"

    print("\n--- DESIGN CONDITIONS ---")
    odb = float(input("Outdoor design dry bulb (°F): "))
    owb = float(input("Outdoor design wet bulb (°F): "))
    idb = float(input("Entering air dry bulb (°F) [80]: ") or 80)
    iewb = float(input("Entering air wet bulb (°F) [67]: ") or 67)

    cap_min = float(input("Min capacity % of load [100]: ") or 100)
    cap_max = float(input("Max capacity % of load [115]: ") or 115)

    print("\nEquipment types:")
    print("  1 = All (A/C single, A/C 2-stage, HP single, HP 2-stage)")
    print("  2 = A/C only")
    print("  3 = Heat Pump only")
    et_choice = input("Choice [1]: ").strip() or "1"
    if et_choice == "2":
        eq_types = [AC_SINGLE, AC_TWO]
    elif et_choice == "3":
        eq_types = [HP_SINGLE, HP_TWO]
    else:
        eq_types = None

    print("\n--- ZONES ---")
    n_zones = int(input("Number of zones: "))
    zones = []
    for i in range(n_zones):
        print(f"\nZone {i+1}:")
        name = input(f"  Zone name/tag [Zone {i+1}]: ").strip() or f"Zone {i+1}"
        tc = float(input(f"  Total cooling load (kBtu/h): "))
        shc = float(input(f"  Sensible cooling load (kBtu/h): "))
        htg_str = input(f"  Heating load (kBtu/h) [0 = skip]: ").strip() or "0"
        htg = float(htg_str)
        zones.append({"name": name, "total_cooling_kbtu": tc,
                       "sensible_cooling_kbtu": shc, "total_heating_kbtu": htg})

    print("\nRunning selection...")
    results = select_equipment(zones, odb, owb, idb, iewb, cap_min, cap_max,
                                equipment_types=eq_types)

    # Print summary
    print("\n" + "="*60)
    print("  SELECTION RESULTS")
    print("="*60)
    for res in results:
        sel = res["selected"]
        if sel is None:
            print(f"\n  {res['zone']}: NO IN-BOUNDS MATCH")
            print(f"    Load: {res['tc_load']} kBtu/h total, {res['shc_load']} kBtu/h sensible")
            ns = res.get("next_smaller"); nl = res.get("next_larger")
            if ns: print(f"    Next smaller: {ns['odu_model']} — {ns['tc_sys']:.1f} kBtu/h ({ns['tc_sys']/res['tc_load']*100:.1f}%)")
            if nl: print(f"    Next larger:  {nl['odu_model']} — {nl['tc_sys']:.1f} kBtu/h ({nl['tc_sys']/res['tc_load']*100:.1f}%)")
            if not ns and not nl: print("    Check: load outside all available equipment range")
        else:
            print(f"\n  {res['zone']}:")
            print(f"    ODU:  {sel['odu_model']} ({sel['odu_series']}, {sel['nom_tons']}T)")
            print(f"    IDU:  {sel['idu_pattern']} ({sel['fc_series']}), {int(sel['cfm'])} CFM")
            print(f"    Cooling:  {sel['tc_sys_kbtu']} kBtu/h total / {sel['shc_sys_kbtu']} kBtu/h sensible")
            print(f"    Load:     {sel['tc_load_kbtu']} / {sel['shc_load_kbtu']} kBtu/h ({sel['tc_pct']}%)")
            if sel['sst_f']:
                print(f"    Op. SST:  {sel['sst_f']}°F  |  SDT: {sel['sdt_f']}°F")
            if sel['eer_site']:
                print(f"    Site EER: {sel['eer_site']} Btu/Wh")
            if sel['htg_cap_kbtu']:
                status = "OK" if sel['htg_ok'] else "⚠ SUPPLEMENTAL HEAT REQUIRED"
                print(f"    HP Htg:   {sel['htg_cap_kbtu']} kBtu/h  {status}")
            n_alt = len(res["all_candidates"])
            if n_alt > 1:
                print(f"    ({n_alt-1} alternate(s) also qualify — see Excel output)")

    # Output files
    out_base = project.replace(" ", "_")
    xlsx_path = f"C:/Users/ghost/adicot/equipment_selection_calculator/{out_base}_schedule.xlsx"
    dxf_path  = f"C:/Users/ghost/adicot/equipment_selection_calculator/{out_base}_schedule.dxf"

    write_excel_schedule(results, xlsx_path, project, odb, owb)
    write_dxf_schedule(results, dxf_path, project, odb, owb)

    return xlsx_path, dxf_path, results


def run_from_config(config_path: str):
    with open(config_path) as f:
        cfg = json.load(f)

    results = select_equipment(
        zones=cfg["zones"],
        outdoor_db=cfg["outdoor_db"],
        outdoor_wb=cfg["outdoor_wb"],
        entering_db=cfg.get("entering_db", 80.0),
        entering_ewb=cfg.get("entering_ewb", 67.0),
        cap_min_pct=cfg.get("cap_min_pct", 100.0),
        cap_max_pct=cfg.get("cap_max_pct", 115.0),
        equipment_types=cfg.get("equipment_types", None),
    )

    project = cfg.get("project_name", "PROJECT")
    out_base = project.replace(" ", "_")
    xlsx_path = cfg.get("output_xlsx", f"C:/Users/ghost/adicot/equipment_selection_calculator/{out_base}_schedule.xlsx")
    dxf_path  = cfg.get("output_dxf",  f"C:/Users/ghost/adicot/equipment_selection_calculator/{out_base}_schedule.dxf")

    write_excel_schedule(results, xlsx_path, project, cfg["outdoor_db"], cfg["outdoor_wb"],
                         cfg.get("cap_min_pct", 100), cfg.get("cap_max_pct", 115))
    write_dxf_schedule(results, dxf_path, project, cfg["outdoor_db"], cfg["outdoor_wb"])
    return xlsx_path, dxf_path, results


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="HVAC Equipment Selector — Adicot Engineering")
    parser.add_argument("--config", type=str, help="Path to JSON config file")
    args = parser.parse_args()

    if args.config:
        run_from_config(args.config)
    else:
        run_interactive()