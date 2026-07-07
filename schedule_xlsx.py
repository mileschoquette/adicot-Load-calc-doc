"""Excel (.xlsx) renderers for the three signed schedules.

These mirror the ReportLab PDFs in hvac_pipeline.py, but as real spreadsheets.
Design Master → compute() → ComputedReport is the shared input; these functions
take the same `computed`/`report`/`config`/`engineer`/`firm` objects the PDF
builders do and write an .xlsx with proper print setup (US Letter, portrait,
fit-to-width, 0.5in margins, repeating header rows).

The xlsx is then rendered to PDF by xlsx_to_pdf.convert() so the delivered PDF
is spreadsheet-origin — which imports cleanly into AutoCAD. See build_all_pdfs.

Pure logic: no Flask, no network. Attributes are read duck-typed off the
ComputedReport dataclasses, so this module does NOT import hvac_pipeline at
module load (avoids a circular import); the few shared helpers it needs are
imported lazily inside build_load_summary_xlsx.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils import get_column_letter

# ── Shared styling ──────────────────────────────────────────────────────
# "Arial" renders as Liberation Sans under headless LibreOffice (metric-
# compatible), matching the Helvetica look of the ReportLab PDFs.
_FONT = "Arial"
TITLE_FONT = Font(name=_FONT, size=16, bold=True)
SUB_FONT   = Font(name=_FONT, size=11)
HDR_FONT   = Font(name=_FONT, size=9, bold=True)
BODY_FONT  = Font(name=_FONT, size=9)
LABEL_FONT = Font(name=_FONT, size=10, bold=True)
VALUE_FONT = Font(name=_FONT, size=10)

CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_NW = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT     = Alignment(horizontal="right", vertical="center")
LEFT_NW   = Alignment(horizontal="left", vertical="center")

_thin = Side(style="thin", color="000000")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

INT_FMT  = "#,##0"
DEC1_FMT = "0.#"


def _finish_print(ws, ncols: int, last_row: int,
                  title_rows: str | None = None,
                  landscape: bool = False) -> None:
    """Apply the common print setup used by all three schedules."""
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = 1  # US Letter
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # as many pages tall as needed
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5,
                                  header=0.2, footer=0.2)
    ws.print_area = f"A1:{get_column_letter(ncols)}{last_row}"
    ws.print_options.horizontalCentered = True
    if title_rows:
        ws.print_title_rows = title_rows


def _box(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    """Apply a thin border to every cell in a rectangular range."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = BORDER


def _int_or_dash(v):
    """Return an int for a real number, else '-' (matches _fmt_int display)."""
    if v is None or v == "":
        return "-"
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return str(v)


def _num_cell(ws, row, col, value):
    """Write a right-aligned integer cell, or a plain '-' string for a dash."""
    v = _int_or_dash(value)
    cell = ws.cell(row, col, v)
    cell.alignment = RIGHT
    if isinstance(v, (int, float)):
        cell.number_format = INT_FMT
    return cell


# ══════════════════════════════════════════════════════════════════════
# Ventilation Schedule
# ══════════════════════════════════════════════════════════════════════
def build_ventilation_schedule_xlsx(computed, out_path: Path,
                                    project_name: str | None = None) -> Path:
    out_path = Path(out_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventilation Schedule"
    ncols = 7

    # Title + mechanical-code subtitle
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, "VENTILATION SCHEDULE"); t.font = TITLE_FONT; t.alignment = CENTER_NW
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(2, 1, computed.mechanical_code or ""); s.font = SUB_FONT; s.alignment = CENTER_NW
    ws.row_dimensions[1].height = 22

    # Header block: rows 4-7
    hr = 4
    ws.merge_cells(start_row=hr, start_column=1, end_row=hr + 3, end_column=1)   # Room
    ws.merge_cells(start_row=hr, start_column=2, end_row=hr + 3, end_column=2)   # Room Type
    ws.merge_cells(start_row=hr, start_column=3, end_row=hr, end_column=4)       # OA, Occupants
    ws.merge_cells(start_row=hr, start_column=5, end_row=hr, end_column=6)       # OA, Area
    ws.merge_cells(start_row=hr, start_column=7, end_row=hr + 1, end_column=7)   # Ventilation Rate

    ws.cell(hr, 1, "Room"); ws.cell(hr, 2, "Room Type")
    ws.cell(hr, 3, "Outdoor Air, Occupants"); ws.cell(hr, 5, "Outdoor Air, Area")
    ws.cell(hr, 7, "Ventilation Rate")
    ws.cell(hr + 1, 3, "Rate"); ws.cell(hr + 1, 4, "People")
    ws.cell(hr + 1, 5, "Rate"); ws.cell(hr + 1, 6, "Area")
    ws.cell(hr + 2, 3, "[CFM/person]"); ws.cell(hr + 2, 5, "[cfm/ft²]")
    ws.cell(hr + 2, 6, "[ft²]"); ws.cell(hr + 2, 7, "[CFM]")
    ws.cell(hr + 3, 3, "Rp"); ws.cell(hr + 3, 4, "Pz")
    ws.cell(hr + 3, 5, "Ra"); ws.cell(hr + 3, 6, "Az"); ws.cell(hr + 3, 7, "Vbz*")

    for r in range(hr, hr + 4):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.font = HDR_FONT
            cell.alignment = LEFT if (c <= 2 and r == hr) else CENTER
    _box(ws, hr, 1, hr + 3, ncols)

    # Data rows
    row = hr + 4
    total_people = 0
    for rm in computed.rooms:
        ws.cell(row, 1, rm.room).alignment = LEFT
        ws.cell(row, 2, rm.room_type).alignment = LEFT
        c3 = ws.cell(row, 3, float(rm.rp_cfm_per_person) if rm.rp_cfm_per_person else 0)
        c3.number_format = DEC1_FMT; c3.alignment = RIGHT
        c4 = ws.cell(row, 4, int(rm.pz_people) if rm.pz_people else 0)
        c4.number_format = INT_FMT; c4.alignment = RIGHT
        # Ra display string: "0.06", "2 ACH", or "0"
        ws.cell(row, 5, rm.ra_display or "0").alignment = RIGHT
        c6 = ws.cell(row, 6, int(round(rm.area_ft2 or 0))); c6.number_format = INT_FMT; c6.alignment = RIGHT
        c7 = ws.cell(row, 7, int(round(rm.vent_cfm or 0))); c7.number_format = INT_FMT; c7.alignment = RIGHT
        for c in range(1, ncols + 1):
            ws.cell(row, c).font = BODY_FONT
        if rm.pz_people:
            total_people += int(rm.pz_people)
        row += 1

    _box(ws, hr + 4, 1, row - 1, ncols)

    # Footer: occupants (left) + total OA (right)
    total_oa = int(round(computed.total_vent_oa_cfm or 0))
    fr = row
    ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=3)
    fc = ws.cell(fr, 1, f"{total_people} Occupants"); fc.font = BODY_FONT; fc.alignment = LEFT_NW
    ws.merge_cells(start_row=fr, start_column=4, end_row=fr, end_column=ncols)
    ft = ws.cell(fr, 4, f"Total Min. OA {total_oa} CFM"); ft.font = HDR_FONT; ft.alignment = RIGHT
    _box(ws, fr, 1, fr, ncols)

    # Column widths (chars ≈ the PDF inch proportions)
    for i, w in enumerate([18, 32, 12, 8, 12, 9, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _finish_print(ws, ncols, fr, title_rows=f"{hr}:{hr + 3}")
    wb.save(out_path)
    return out_path


# ══════════════════════════════════════════════════════════════════════
# Building Air Balance
# ══════════════════════════════════════════════════════════════════════
def build_air_balance_xlsx(computed, out_path: Path,
                           project_name: str | None = None,
                           config=None) -> Path:
    out_path = Path(out_path)
    show_imc_footnote = getattr(config, "bldg_exhaust_all_toilet", False) if config else False
    wb = Workbook()
    ws = wb.active
    ws.title = "Air Balance"
    ncols = 6

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, "Building Air Balance"); t.font = TITLE_FONT; t.alignment = CENTER_NW
    ws.row_dimensions[1].height = 22

    # Header rows 3-4
    hr = 3
    for c, txt in enumerate(["Zone", "Supply", "Return", "Bldg Ventilation",
                             "Bldg Exhaust", "Air Balance"], start=1):
        ws.cell(hr, c, txt)
    for c in range(2, ncols + 1):
        ws.cell(hr + 1, c, "[cfm]")
    ws.merge_cells(start_row=hr, start_column=1, end_row=hr + 1, end_column=1)  # Zone spans
    for r in range(hr, hr + 2):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.font = HDR_FONT
            cell.alignment = LEFT if c == 1 else CENTER
    _box(ws, hr, 1, hr + 1, ncols)

    # Data rows
    row = hr + 2
    for z in computed.zones:
        ws.cell(row, 1, z.zone_name).alignment = LEFT
        _num_cell(ws, row, 2, z.supply_cfm)
        _num_cell(ws, row, 3, z.return_cfm)
        _num_cell(ws, row, 4, z.vent_oa_cfm)
        _num_cell(ws, row, 5, "-" if z.bldg_exhaust_cfm == 0 else z.bldg_exhaust_cfm)
        _num_cell(ws, row, 6, z.air_balance_cfm)
        for c in range(1, ncols + 1):
            ws.cell(row, c).font = BODY_FONT
        row += 1
    _box(ws, hr + 2, 1, row - 1, ncols)

    # Totals row
    tr = row
    lbl = ws.cell(tr, 3, "Totals:"); lbl.font = HDR_FONT; lbl.alignment = RIGHT
    _num_cell(ws, tr, 4, computed.total_vent_oa_cfm)
    _num_cell(ws, tr, 5, "-" if computed.total_bldg_exhaust_cfm == 0 else computed.total_bldg_exhaust_cfm)
    _num_cell(ws, tr, 6, computed.total_air_balance_cfm)
    for c in (4, 5, 6):
        ws.cell(tr, c).font = HDR_FONT
    _box(ws, tr, 1, tr, ncols)

    last = tr
    if show_imc_footnote:
        nr = tr + 2
        ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=ncols)
        note = ws.cell(nr, 1,
                       "All building exhaust is intermittent toilet/accessory exhaust "
                       "per IMC Table 403.4.2. No continuous exhaust.")
        note.font = BODY_FONT
        note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        _box(ws, nr, 1, nr, ncols)
        ws.row_dimensions[nr].height = 28
        last = nr

    for i, w in enumerate([34, 12, 12, 16, 13, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _finish_print(ws, ncols, last, title_rows=f"{hr}:{hr + 1}")
    wb.save(out_path)
    return out_path


# ══════════════════════════════════════════════════════════════════════
# Load Summary cover sheet
# ══════════════════════════════════════════════════════════════════════
def build_load_summary_xlsx(computed, report, config, engineer, firm,
                            out_path: Path,
                            project_name: str | None = None) -> Path:
    # Lazy import to avoid a circular import at module load.
    from hvac_pipeline import (grains_water_difference, _parse_calc_date,
                               _fmt_date)

    out_path = Path(out_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Load Summary"
    ncols = 6

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, "HEATING AND COOLING LOAD SUMMARY SHEET")
    t.font = TITLE_FONT; t.alignment = CENTER_NW
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(2, 1, computed.energy_code or ""); s.font = SUB_FONT; s.alignment = CENTER_NW

    # Engineer block (label in cols A:B, value in cols C:F)
    lic_label = f"{engineer.state_full} Registered Professional Engineer:"
    eng_rows = [
        ("Calculations Performed by:", engineer.name),
        ("Contact:", f"{engineer.email}   {engineer.phone}"),
        (lic_label, f"Lic. No.: {computed.license_number}"),
        ("Date:", _fmt_date(_parse_calc_date(computed.calc_date))),
    ]
    r = 4
    for label, value in eng_rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        lc = ws.cell(r, 1, label); lc.font = LABEL_FONT; lc.alignment = RIGHT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=ncols)
        vc = ws.cell(r, 3, value); vc.font = VALUE_FONT; vc.alignment = LEFT_NW
        vc.border = Border(bottom=_thin)
        r += 1

    # Project block (bordered)
    rh = f"{int(computed.indoor_rh)}%" if computed.indoor_rh else "-"
    gwd = grains_water_difference(report)
    gwd_str = f"{gwd:.2f} [grains moisture/lb dry air]" if gwd is not None else "-"
    proj_address = (getattr(config, "project_address", None)
                    or computed.project_address
                    or computed.weather_station)

    def _deg(v):
        return f"{int(round(v))}° F" if v else "-"

    proj_rows = [
        ("Project Name", computed.project_name),
        ("Address", proj_address),
        ("Weather Station", computed.weather_station),
        ("Sizing Method", "CLTD"),
        ("Outdoor Dry Bulb", _deg(computed.osa_high_db_f)),
        ("Outdoor Wet Bulb", _deg(computed.osa_high_wb_f)),
        ("Indoor Dry Bulb", _deg(computed.indoor_dry_bulb_f)),
        ("RH", rh),
        ("Grains Water Difference", gwd_str),
    ]
    r += 1
    p_start = r
    for label, value in proj_rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        lc = ws.cell(r, 1, label); lc.font = LABEL_FONT; lc.alignment = RIGHT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=ncols)
        vc = ws.cell(r, 3, str(value)); vc.font = VALUE_FONT; vc.alignment = LEFT_NW
        r += 1
    _box(ws, p_start, 1, r - 1, ncols)

    # Load table
    r += 1
    hr = r
    for c, txt in enumerate(["Zone", "Area [ft²]", "Total Cooling [Btu/h]",
                             "Total Sensible Gain [Btu/h]", "Total Latent Gain [Btu/h]",
                             "Total Heating [Btu/h]"], start=1):
        cell = ws.cell(hr, c, txt); cell.font = HDR_FONT
        cell.alignment = LEFT if c == 1 else CENTER
    _box(ws, hr, 1, hr, ncols)

    r = hr + 1
    for z in computed.zones:
        ws.cell(r, 1, z.zone_name).alignment = LEFT
        _num_cell(ws, r, 2, z.area_ft2)
        _num_cell(ws, r, 3, z.cooling_total_btuh)
        _num_cell(ws, r, 4, z.cooling_sensible_btuh)
        _num_cell(ws, r, 5, z.cooling_latent_btuh)
        _num_cell(ws, r, 6, z.heating_btuh)
        for c in range(1, ncols + 1):
            ws.cell(r, c).font = BODY_FONT
        r += 1
    _box(ws, hr, 1, r - 1, ncols)

    # Firm footer at the bottom of every printed page
    ws.oddFooter.center.text = f"{firm.line1}\n{firm.line2}"
    ws.oddFooter.center.size = 9
    ws.oddFooter.center.font = _FONT

    for i, w in enumerate([26, 12, 16, 18, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _finish_print(ws, ncols, r - 1)
    wb.save(out_path)
    return out_path
