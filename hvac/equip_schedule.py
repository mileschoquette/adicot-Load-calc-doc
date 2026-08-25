"""Cross-engine glue for the Equipment Selection tab.

Owns everything that needs to know about AC/HP selection (hvac_selector.py),
ERV sizing (erv_calculator/), and dehumidifier sizing (dehumid_calc.py) at
once, so those three modules can each stay single-purpose. Two jobs:

1. Compute how a *specific* chosen ERV or dehumidifier changes the sensible/
   total cooling load an AC/HP zone must be sized against (apply_load_deltas),
   so picking different equipment quantities actually changes the resulting
   AC/HP selection.
2. Render the combined AC/HP + ERV + Dehumidification workbook as one sheet.
"""
from __future__ import annotations

from openpyxl import Workbook

from hvac import hvac_selector as sel

# Each engine is independently optional — a missing dependency for one (e.g.
# pandas for dehumid_calc) must not break the other.
try:
    from hvac import dehumid_calc as dh
except Exception:
    dh = None

try:
    from erv_calculator import load_impact, performance as erv_performance
except Exception:
    load_impact = None
    erv_performance = None


def dehumid_adjustment(model_row, units: int, config: dh.DehumidConfig) -> dict:
    """Real load impact of `units` copies of a specific catalog row (a pandas
    Series from dehumid_calc.load_database()), as opposed to
    compute_dehumidification's target-driven 15%/85% split. Returns signed
    BTU/h deltas ready for apply_load_deltas."""
    derated_pints_day = dh.derated_capacity(model_row, config)
    pints_day_removed = derated_pints_day * units
    latent_removed_btuh = (pints_day_removed / 24.0) * dh.LATENT_BTU_PER_PINT
    sensible_added_btuh = dh.sensible_heat_penalty(pints_day_removed)
    return {
        "derated_capacity_pints_day": derated_pints_day,
        "pints_day_removed": pints_day_removed,
        "latent_removed_btuh": latent_removed_btuh,
        "sensible_added_btuh": sensible_added_btuh,
        "sensible_delta_btuh": sensible_added_btuh,
        "total_delta_btuh": sensible_added_btuh - latent_removed_btuh,
    }


def erv_adjustment(cfm: float, t1_f: float, w1_gr: float, t3_f: float, w3_gr: float,
                    perf: "erv_performance.ErvPerformance", airflow_frac: float, season: str) -> dict:
    """Real sensible/latent reduction a specific ERV unit delivers, straight
    from erv_calculator.load_impact.compute_erv_impact. Deltas are negative:
    recovered load no longer has to be handled by the AC/HP."""
    impact = load_impact.compute_erv_impact(cfm, t1_f, w1_gr, t3_f, w3_gr, perf, airflow_frac, season)
    return {
        "impact": impact,
        "sensible_delta_btuh": -impact.sensible_reduction_btuh,
        "total_delta_btuh": -(impact.sensible_reduction_btuh + impact.latent_reduction_btuh),
    }


def apply_load_deltas(tc_kbtu: float, shc_kbtu: float, *adjustments: dict) -> tuple:
    """Adds one or more signed BTU/h {sensible_delta_btuh, total_delta_btuh}
    adjustments onto a zone's tc/shc (kBtu/h). Multiple adjustments (e.g. an
    ERV and a dehumidifier on the same zone) combine additively. Returns
    (adjusted_tc_kbtu, adjusted_shc_kbtu)."""
    sensible_delta_kbtu = sum(a["sensible_delta_btuh"] for a in adjustments) / 1000.0
    total_delta_kbtu = sum(a["total_delta_btuh"] for a in adjustments) / 1000.0
    return tc_kbtu + total_delta_kbtu, shc_kbtu + sensible_delta_kbtu


def _section_title(ws, row: int, title: str, span: str = "A{row}:F{row}") -> None:
    cell_range = span.format(row=row)
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = title
    from openpyxl.styles import Font
    c.font = Font(bold=True, size=11, color="1F4E79")


def _kv_row(ws, row: int, label: str, value) -> None:
    sel.dcell(ws, row, 1, label, bold=False, center_align=False)
    sel.dcell(ws, row, 2, value, center_align=False)


def write_erv_block(ws, start_row: int, erv_result: dict) -> int:
    """Titled key/value section describing the chosen ERV unit and its load
    impact on the zone it serves. Returns the next free row."""
    row = start_row
    _section_title(ws, row, "VENTILATION — ENERGY RECOVERY")
    row += 2

    impact = erv_result["impact"]
    rows = [
        ("Zone served", erv_result["zone"]),
        ("Unit", f"{erv_result.get('manufacturer', '')} {erv_result['model']}".strip()),
        ("Outdoor airflow (CFM)", round(erv_result["cfm"], 0)),
        ("Season / airflow fraction", f"{erv_result['season']} @ {erv_result['airflow_frac']:.2f}"),
        ("Raw OA sensible / latent (BTU/h)",
         f"{impact.raw_sensible_btuh:,.0f} / {impact.raw_latent_btuh:,.0f}"),
        ("Net OA sensible / latent after recovery (BTU/h)",
         f"{impact.net_sensible_btuh:,.0f} / {impact.net_latent_btuh:,.0f}"),
        ("Sensible / latent recovered (BTU/h)",
         f"{impact.sensible_reduction_btuh:,.0f} / {impact.latent_reduction_btuh:,.0f}"),
        ("AC/HP sensible load adjustment (BTU/h)", f"{erv_result['sensible_delta_btuh']:+,.0f}"),
        ("AC/HP total load adjustment (BTU/h)", f"{erv_result['total_delta_btuh']:+,.0f}"),
    ]
    if erv_result.get("frost_risk"):
        rows.append(("Frost risk", "YES — outdoor design temp below core frost threshold; verify defrost strategy"))

    for label, value in rows:
        _kv_row(ws, row, label, value)
        row += 1
    return row + 1


def write_dehumid_block(ws, start_row: int, dehumid_result: dict) -> int:
    """Titled key/value section describing the chosen dehumidifier(s) and
    their load impact on the zone they serve. Returns the next free row."""
    row = start_row
    _section_title(ws, row, "DEHUMIDIFICATION")
    row += 2

    qty = dehumid_result["units"]
    rows = [
        ("Zone served", dehumid_result["zone"]),
        ("Unit", f"{qty}x {dehumid_result.get('manufacturer', '')} {dehumid_result['model']}".strip()),
        ("Derated capacity (pints/day, each)", round(dehumid_result["derated_capacity_pints_day"], 0)),
        ("Total moisture removed (pints/day)", round(dehumid_result["pints_day_removed"], 0)),
        ("Latent load removed (BTU/h)", f"{dehumid_result['latent_removed_btuh']:,.0f}"),
        ("Sensible heat added back (BTU/h)", f"{dehumid_result['sensible_added_btuh']:,.0f}"),
        ("AC/HP sensible load adjustment (BTU/h)", f"{dehumid_result['sensible_delta_btuh']:+,.0f}"),
        ("AC/HP total load adjustment (BTU/h)", f"{dehumid_result['total_delta_btuh']:+,.0f}"),
    ]
    for label, value in rows:
        _kv_row(ws, row, label, value)
        row += 1
    return row + 1


def write_combined_schedule(ac_hp_results: list, erv_result: dict | None, dehumid_result: dict | None,
                             filepath: str, project_name: str = "PROJECT",
                             outdoor_db: float = None, outdoor_wb: float = None,
                             cap_min_pct: float = 100, cap_max_pct: float = 115) -> None:
    """Builds one workbook, one continuous sheet: title -> design conditions ->
    AC/HP block -> ERV block (if included) -> Dehumidification block (if
    included) -> notes. Saves once."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Schedule"

    cond_str = ""
    if outdoor_db: cond_str += f"  Outdoor DB: {outdoor_db}°F"
    if outdoor_wb: cond_str += f"  |  Outdoor WB: {outdoor_wb}°F"
    cond_str += "  |  Capacities at design conditions (corrected)"
    sel.write_title_block(ws, f"MECHANICAL — EQUIPMENT SCHEDULE — {project_name}", cond_str)

    row = sel.write_ac_hp_block(ws, 3, ac_hp_results)
    row += 1

    notes = list(sel.ac_hp_notes(cap_min_pct, cap_max_pct))

    if erv_result is not None or dehumid_result is not None:
        adjusted_zones = sorted({r["zone"] for r in (erv_result, dehumid_result) if r})
        notes.append(
            f"9. AC/HP sizing for {', '.join(adjusted_zones)} reflects the ERV/dehumidifier load "
            "adjustments below — it is NOT sized against the raw, un-adjusted Design Master load."
        )

    if erv_result is not None:
        row = write_erv_block(ws, row, erv_result)

    if dehumid_result is not None:
        row = write_dehumid_block(ws, row, dehumid_result)
        notes.extend(dh.DEFAULT_CAVEATS)

    sel.write_notes_section(ws, row, notes)

    wb.save(filepath)
    print(f"Combined schedule saved: {filepath}")
